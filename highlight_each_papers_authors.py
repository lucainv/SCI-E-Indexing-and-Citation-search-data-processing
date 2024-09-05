import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def standardize_author_name(author_name):
    """
    将作者姓名标准化为 '姓, 名' 或 '姓, 名1名2' 的形式，确保仅姓首字母大写。
    如果姓名中有短横线，则去掉短横线。
    """
    # 去掉短横线
    author_name = author_name.replace('-', ' ')

    parts = author_name.split()

    if len(parts) == 1:
        # 只有姓或名
        surname = parts[0].capitalize()
        return surname
    elif len(parts) > 1:
        # 包含姓和名
        surname = parts[0].capitalize()
        given_name = ''.join(parts[1:]).capitalize()
        return f"{surname}, {given_name}"


def highlight_name(input_file, output_file, column_name, names):
    """
    高亮指定列中的名字，并同时高亮'Authors'列，保留原始作者姓名格式。
    """
    # 读取 .xls 文件并转换为 .xlsx 文件
    intermediate_file = 'intermediate.xlsx'

    # 使用 pandas 读取 .xls 文件
    df = pd.read_excel(input_file)

    # 将数据保存为 .xlsx 文件
    df.to_excel(intermediate_file, index=False)

    # 使用 openpyxl 加载 .xlsx 文件
    wb = load_workbook(intermediate_file)
    ws = wb.active

    # 设置黄色填充
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 获取指定列的索引
    col_idx = df.columns.get_loc(column_name) + 1
    author_col_idx = df.columns.get_loc('Authors') + 1  # 获取'Authors'列的索引

    # 统计总数据条数和未被高亮的数据条数
    total_count = 0
    non_highlight_count = 0
    highlight_count = 0

    # 遍历指定列，查找包含搜索字符串的单元格并设置高亮
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        cell = row[col_idx - 1]  # 列的索引从0开始，因此需要减去1
        author_cell = row[author_col_idx - 1]  # 同上

        total_count += 1
        cell_values_standardized = [standardize_author_name(value.strip()) for value in str(cell.value).split(';')]
        flag = 0
        for name in names:
            if name in cell_values_standardized:
                cell.fill = yellow_fill
                author_cell.fill = yellow_fill  # 同时高亮'Authors'字段的单元格
                flag = 1
                highlight_count += 1
                break
        if flag == 0:
            non_highlight_count += 1

    # 保存修改后的 Excel 文件
    wb.save(output_file)

    return total_count, highlight_count, non_highlight_count


# 定义文件路径
base_folder = r'C:\Users\Lenovo\pythonProject'
input_folder = os.path.join(base_folder, 'examples', '张健示例', 'SCI-E引用数据')
output_folder = os.path.join(base_folder, 'data_output')
input_file = os.path.join(input_folder, 'SCI-E引用格式.txt')
output_file = os.path.join(output_folder, 'qingdan.xlsx')
papers_file = os.path.join(base_folder, 'examples', '张健示例', 'papers.xlsx')  # Corrected path

# 确保输出目录存在
os.makedirs(output_folder, exist_ok=True)

# 初始化计数器和列表
paper_sequence = 0
citation_sequence = 0
paper_list = []
citation_list = []
file_names = []
self_cited_authors = []

# 定义需要跳过引用计数的字符串集合
skip_strings = {'N/A', '无引用', 'n/a', 'NA'}

# 标志位用于跟踪何时增加序列
new_paper_flag = True
new_citation_flag = True

# 读取论文文件以将序列号映射到作者
papers_df = pd.read_excel(papers_file)  # Removed engine='xlrd'
sequence_to_authors = dict(zip(papers_df['论文清单序号'], papers_df['Author Full Names']))

# 读取 txt 文件
with open(input_file, 'r', encoding='utf-8') as file:
    for line in file:
        line_content = line.strip()

        if line_content:  # 非空行
            paper_num = None
            citation_num = None
            file_name = None
            authors = None

            # 处理论文序号
            if new_paper_flag:
                paper_sequence += 1
                paper_num = paper_sequence
                new_paper_flag = False

            # 处理引用序号
            if line_content not in skip_strings:
                if new_citation_flag:
                    citation_num = citation_sequence
                    citation_sequence += 1
                    new_citation_flag = False
                    # 根据引用序号确定文件名
                    file_name = f'savedrecs ({citation_num}).xls' if citation_num > 0 else 'savedrecs.xls'

            # 仅在引用序号不为空时，根据论文序号检索自引作者
            if citation_num is not None and paper_num in sequence_to_authors:
                authors = sequence_to_authors[paper_num]
            else:
                authors = None  # 确保在引用序号为空时 authors 为 None

            # 仅在至少一个序列号已分配时才添加到列表中
            if paper_num is not None or citation_num is not None:
                paper_list.append(paper_num)
                citation_list.append(citation_num)
                file_names.append(file_name)
                self_cited_authors.append(authors)

        else:  # 空行
            new_paper_flag = True  # 设置标志位表示可以开始一个新论文
            new_citation_flag = True  # 设置标志位表示可以开始一个新引用

# 创建一个 DataFrame 并将其保存为 Excel 文件
df = pd.DataFrame({
    '论文清单序号': paper_list,
    '有引用论文序号': citation_list,
    '有引用论文的文件名': file_names,
    '自引作者清单': self_cited_authors
})

df.to_excel(output_file, index=False)

# 读取论文文件
papers_df = pd.read_excel(papers_file)  # Removed engine='xlrd'

# 初始化总计数器
total_count_sum = 0
highlight_count_sum = 0
non_highlight_count_sum = 0

# 处理每一行以扩展和格式化自引作者
for index, row in df.iterrows():
    citation_num = row['有引用论文序号']
    file_name = row['有引用论文的文件名']
    authors = row['自引作者清单']

    if pd.notna(citation_num) and file_name and pd.notna(authors):
        # 分割并标准化自引作者清单中的作者姓名
        author_list = [standardize_author_name(author.strip()) for author in authors.split('; ')]

        # 定义用于高亮的路径
        file_path = os.path.join(input_folder, file_name)
        highlighted_file_path = os.path.join(output_folder, f'{os.path.splitext(file_name)[0]}_highlighted.xlsx')

        # 高亮匹配的单元格
        total_count, highlight_count, non_highlight_count = highlight_name(file_path, highlighted_file_path,
                                                                           'Author Full Names', author_list)

        # 更新 DataFrame 的计数
        df.loc[index, '总被引数'] = total_count
        df.loc[index, '自引数'] = highlight_count
        df.loc[index, '他引数'] = non_highlight_count

        # 更新总计数
        total_count_sum += total_count
        highlight_count_sum += highlight_count
        non_highlight_count_sum += non_highlight_count

# 保存更新后的 DataFrame 到 Excel
df.to_excel(output_file, index=False)

# 打印总计数
print(f'总被引数: {total_count_sum}')
print(f'自引数: {highlight_count_sum}')
print(f'他引数: {non_highlight_count_sum}')

print("Highlighting and saving complete.")
