import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from copy import copy  # 导入copy函数

def copy_cell_style(src_cell, dest_cell):
    # Set font to Arial
    dest_cell.font = Font(name='Arial')

    # Set alignment to left
    dest_cell.alignment = Alignment(horizontal='left')

    # Set border to None
    dest_cell.border = Border(left=Side(border_style=None), right=Side(border_style=None),
                              top=Side(border_style=None), bottom=Side(border_style=None))

    # Copy the fill
    dest_cell.fill = copy(src_cell.fill)

    # Copy the number format
    dest_cell.number_format = src_cell.number_format

    # Copy protection
    dest_cell.protection = copy(src_cell.protection)

    # Copy alignment
    dest_cell.alignment = copy(src_cell.alignment)

# 读取 citation_output.xlsx 文件
citation_file_path = 'data_output/citation_output.xlsx'
citation_wb = load_workbook(citation_file_path)
citation_ws = citation_wb.active

# 获取 citation_output.xlsx 文件的内容
citation_df = pd.read_excel(citation_file_path, header=None)  # 没有列名

# 初始化 new_number 和统计数据
new_number = 0
stats_data = []

# 获取 A 列和 B 列的数据
a_column = citation_df[0]
b_column = citation_df[1]

# 定义黄色填充的颜色索引
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 遍历 citation_output.xlsx 文件的行
row_offset = 0
for index in range(len(citation_df)):
    old_number = a_column[index]  # 被引文献序号
    b_column_value = b_column[index]  # B列值

    total_citations = 0
    self_citations = 0
    external_citations = 0

    if pd.notna(old_number) and b_column_value != '无引用':
        # 查找下一个被引文献的位置
        next_citation_index = index + 1
        while next_citation_index < len(citation_df) and pd.isna(a_column[next_citation_index]):
            next_citation_index += 1

        # 确定插入的 savedrecs 文件路径
        savedrecs_file_name = 'savedrecs_highlighted.xlsx' if new_number == 0 else f'savedrecs ({new_number})_highlighted.xlsx'
        savedrecs_file_path = f'data_output/{savedrecs_file_name}'

        # 读取 savedrecs 文件内容
        savedrecs_wb = load_workbook(savedrecs_file_path)
        savedrecs_ws = savedrecs_wb.active

        # 将生成器转换为列表以获取行数
        savedrecs_rows = list(savedrecs_ws.iter_rows(values_only=False))
        savedrecs_row_count = len(savedrecs_rows)
        total_citations = savedrecs_row_count - 1  # 总被引数

        # 计算自引数
        for row in savedrecs_rows[1:]:
            cell = row[5]  # 假设 F 列是第 6 列，索引为 5
            if cell.fill.start_color.index == yellow_fill.start_color.index:  # 检查填充颜色
                self_citations += 1

        external_citations = total_citations - self_citations  # 他引数

        # 确定插入位置
        if next_citation_index < len(citation_df):
            insert_position = next_citation_index + row_offset + 1  # 下一个被引文献的前一行
        else:
            # 找到第一个空行位置
            insert_position = len(citation_df) + row_offset + 1

        # 复制 savedrecs 文件内容到插入位置
        for i, row in enumerate(savedrecs_rows):
            citation_ws.insert_rows(insert_position + i)  # 插入 savedrecs 文件内容
            for j, cell in enumerate(row):
                new_cell = citation_ws.cell(row=insert_position + i, column=j + 1, value=cell.value)
                copy_cell_style(cell, new_cell)  # 复制单元格样式

        # Format the inserted rows
        for row in citation_ws.iter_rows(min_row=insert_position, max_row=insert_position + savedrecs_row_count - 1,
                                         min_col=1, max_col=citation_ws.max_column):
            for cell in row:
                # Ensure Arial font, left alignment, and no borders
                cell.font = Font(name='Arial')
                cell.alignment = Alignment(horizontal='left')
                cell.border = Border(left=Side(border_style=None), right=Side(border_style=None),
                                     top=Side(border_style=None), bottom=Side(border_style=None))

        # 在插入内容的下方再插入两行空行
        citation_ws.insert_rows(insert_position + savedrecs_row_count, 2)

        row_offset += savedrecs_row_count + 2  # 增加偏移量
        new_number += 1  # 更新 new_number
        index = next_citation_index  # 跳到下一个被引文献的位置

    else:
        total_citations = 0
        self_citations = 0
        external_citations = 0

    # 记录统计数据
    if pd.notna(old_number):
        stats_data.append([old_number, total_citations, self_citations, external_citations])

# 创建统计 DataFrame 并保存为 count.xlsx
stats_df = pd.DataFrame(stats_data, columns=['被引文献序号', '总被引数', '自引数', '他引数'])
stats_df = stats_df[stats_df['被引文献序号'].notna()]  # 删除被引文献下方A列为空的行
stats_df.to_excel('data_output/4_SCI-E引用统计表.xlsx', index=False)

# 保存最终结果为 citation_papers.xlsx 文件
citation_wb.save('data_output/3_SCI-E引用明细表.xlsx')

# Load citation statistics from count.xlsx
count_df = pd.read_excel('data_output/4_SCI-E引用统计表.xlsx')

# Load citation_for_word.xlsx into a DataFrame
citation_for_word_path = 'data_output/citation_for_word.xlsx'
citation_for_word_df = pd.read_excel(citation_for_word_path, header=None)

# Create new columns in DataFrame for total citations and external citations
citation_for_word_df['总被引数'] = None
citation_for_word_df['他引数'] = None

# Create a dictionary for quick lookup of statistics
stats_dict = count_df.set_index('被引文献序号').to_dict(orient='index')

# Fill the new columns with data from count.xlsx
for idx, row in citation_for_word_df.iterrows():
    citation_number = row[0]
    if citation_number in stats_dict:
        citation_stats = stats_dict[citation_number]
        citation_for_word_df.at[idx, '总被引数'] = citation_stats['总被引数']
        citation_for_word_df.at[idx, '他引数'] = citation_stats['他引数']

# Remove the first four characters from column A
citation_for_word_df[0] = citation_for_word_df[0].astype(str).str[4:]

# Remove rows where all cells are empty
citation_for_word_df = citation_for_word_df.dropna(how='all').reset_index(drop=True)

# Save updated DataFrame to a new Excel file
updated_file_path = 'data_output/citations_count_all.xlsx'
with pd.ExcelWriter(updated_file_path, engine='openpyxl') as writer:
    citation_for_word_df.to_excel(writer, index=False, header=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Apply Arial font to all cells and remove borders
    arial_font = Font(name='Arial')
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = arial_font
            cell.alignment = Alignment(horizontal='left')
            cell.border = Border(left=Side(border_style=None), right=Side(border_style=None),
                                 top=Side(border_style=None), bottom=Side(border_style=None))

print("Data has been processed, saved, and formatted.")

# Load the cleaned Excel file
wb = load_workbook(filename='data_output/citations_count_all.xlsx')
ws = wb.active  # Assuming we want to modify the active sheet

# Define a function to check if a row is empty
def is_row_empty(row):
    for cell in row:
        if cell.value is not None and str(cell.value).strip():
            return False
    return True

# Reverse iterate over rows to avoid index interference when deleting
rows_to_delete = []
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
    if is_row_empty(row):
        rows_to_delete.append(idx)

# Delete rows
for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

# Save the cleaned file
wb.save('data_output/5_SCI-E引用格式表_for_word.xlsx')  # Save as new file to avoid overwriting original

print("Empty rows have been removed and the file has been saved.")