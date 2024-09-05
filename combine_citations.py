import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def convert_txt_to_xlsx(txt_file_path, xlsx_file_path):
    with open(txt_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    data = []
    row_num = 1
    new_row = True

    for line in lines:
        stripped_line = line.strip()
        if stripped_line == "":
            new_row = True
        else:
            if new_row:
                row_data = [f"被引文献{row_num}"] + stripped_line.split('\t')
                row_num += 1
                new_row = False
            else:
                row_data = [""] + stripped_line.split('\t')
            data.append(row_data)

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Replace unwanted values with "无引用"
    replacements = ["NA", "n/a", "N/A", "无", "-", "——"]
    df.replace(replacements, "无引用", inplace=True)

    # Write to Excel file
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Apply bold font to cells with content
        bold_font = Font(bold=True, name='微软雅黑')
        normal_font = Font(name='微软雅黑')
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = bold_font if cell.value else normal_font


def process_xlsx(input_file_path, output_file_path):
    wb = load_workbook(input_file_path)
    ws = wb.active

    # Delete specific columns: D, G, I, J
    columns_to_delete = [4, 7, 9, 10]  # 1-based index for columns
    for col in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col)

    max_row = ws.max_row
    col_count = ws.max_column

    # To store merged cell ranges for column A
    merged_ranges = []

    # Process column A
    start_row = 1
    while start_row <= max_row:
        cell_value = ws.cell(row=start_row, column=1).value
        if cell_value:
            end_row = start_row
            while end_row + 1 <= max_row and ws.cell(row=end_row + 1, column=1).value is None:
                end_row += 1
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            merged_ranges.append((start_row, end_row))
            start_row = end_row + 1
        else:
            start_row += 1

    # Function to concatenate cell values with newline characters
    def concatenate_values(start_row, end_row, col):
        values = []
        for row in range(start_row, end_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                values.append(cell_value)
        return "\n".join(values)

    # Process column B: Remove content after the first semicolon
    for row in range(1, max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value:
            ws.cell(row=row, column=2).value = cell_value.split(';')[0].strip()

    # Replace empty cells in columns B, C, D, E, F with "/"
    for col in range(2, col_count + 1):
        if col in {2, 3, 4, 5, 6}:  # B, C, D, E, F columns
            for row in range(1, max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None or cell_value == "":
                    ws.cell(row=row, column=col).value = '/'

    # Merge cells based on column A merged ranges
    for col in range(1, col_count + 1):
        for start_row, end_row in merged_ranges:
            merged_value = concatenate_values(start_row, end_row, col)
            ws.cell(row=start_row, column=col).value = merged_value
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    # Save the processed file
    wb.save(output_file_path)


# Use the functions to process files
txt_file_path = 'C:/Users/Lenovo/pythonProject/examples/张健示例/SCI-E引用数据/SCI-E引用格式.txt'
xlsx_file_path = 'C:/Users/Lenovo/pythonProject/data_output/citation_output.xlsx'
convert_txt_to_xlsx(txt_file_path, xlsx_file_path)

output_file_path = 'C:/Users/Lenovo/pythonProject/data_output/citation_for_word.xlsx'
process_xlsx(xlsx_file_path, output_file_path)

print("Data has been processed and saved.")
