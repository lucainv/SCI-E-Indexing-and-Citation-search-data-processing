import os
import re
from docx import Document
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def extract_references_from_docx(docx_path):
    """
    从 Word 文档中提取文献的详细信息，包括作者、题名、刊名、年、卷期页码、开始页和结束页或文献号码。
    """
    if not os.path.exists(docx_path):
        raise FileNotFoundError(f"文件不存在: {docx_path}")

    doc = Document(docx_path)
    references = []

    paper_index = 1

    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue

        parts = [p.strip() for p in text.split('.') if p.strip()]

        if len(parts) < 4:
            continue

        author, title, journal, year_issue_page = parts[0:4]

        if ',' in year_issue_page:
            year = year_issue_page.split(',')[0].strip()
        else:
            year = ""

        page_info = parts[-1]
        start_page, end_page, article_number = None, None, None

        page_match = re.search(r'(\d+)[-—](\d+)', page_info)
        if page_match:
            start_page, end_page = page_match.groups()
        else:
            if ':' in page_info:
                potential_info = page_info.split(':')[-1].strip()
                if not re.search(r'[-—]', potential_info):
                    article_number = potential_info

        reference = {
            '论文清单序号': paper_index,
            '作者': author,
            '题名': title,
            '刊名': journal,
            '年': year,
            '年卷期页码': year_issue_page,
            '开始页': start_page,
            '结束页': end_page,
            '文献号码': article_number
        }
        references.append(reference)

        paper_index += 1

    return references


def export_references_to_excel(references, output_path):
    """
    将提取的文献信息导出为 Excel 格式。
    """
    df = pd.DataFrame(references)

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='张健论文清单')

        workbook = writer.book
        worksheet = writer.sheets['张健论文清单']

        cell_format = workbook.add_format({'font_name': 'Arial'})
        worksheet.set_column('A:I', None, cell_format)

    print(f"文献信息已成功导出至: {output_path}")


def clean_title(title):
    """
    清除标题中的特殊符号，并移除所有不可见字符
    """
    title = re.sub(r'[^\w\s]', '', title)  # 移除标点符号
    title = re.sub(r'\s+', ' ', title)  # 替换多个空格为单个空格
    return title.lower().strip()  # 转为小写并去除两端空格


def match_references(sci_df, references_df):
    """
    匹配文献标题或页码，并添加论文清单序号列。
    """
    sci_df.insert(0, '论文清单序号', None)

    # Convert references_df to dictionaries for faster lookup
    reference_dict = {}
    cleaned_ref_dict = {}

    for _, row in references_df.iterrows():
        ref_title = row['题名'].strip().lower()
        cleaned_ref_title = clean_title(row['题名'])
        reference_dict[ref_title] = row['论文清单序号']
        cleaned_ref_dict[cleaned_ref_title] = row['论文清单序号']

    for i, sci_row in sci_df.iterrows():
        sci_title = sci_row['Article Title'].strip().lower()

        # 将所有相关字段转换为字符串，同时去掉小数点
        sci_start_page = str(int(float(sci_row.get('Start Page', '')))) if str(sci_row.get('Start Page', '')).replace(
            '.', '', 1).isdigit() else str(sci_row.get('Start Page', '')).strip()
        sci_end_page = str(int(float(sci_row.get('End Page', '')))) if str(sci_row.get('End Page', '')).replace('.', '',
                                                                                                                1).isdigit() else str(
            sci_row.get('End Page', '')).strip()
        sci_article_number = str(int(float(sci_row.get('Article Number', '')))) if str(
            sci_row.get('Article Number', '')).replace('.', '', 1).isdigit() else str(
            sci_row.get('Article Number', '')).strip()

        matched = False

        # First, try exact title match
        if sci_title in reference_dict:
            sci_df.at[i, '论文清单序号'] = reference_dict[sci_title]
            continue

        # Second, try cleaned title match
        cleaned_sci_title = clean_title(sci_row['Article Title']).strip()
        if cleaned_sci_title in cleaned_ref_dict:
            sci_df.at[i, '论文清单序号'] = cleaned_ref_dict[cleaned_sci_title]
            continue

        # Third, try title first 5 words match
        sci_title_first_five_words = ' '.join(sci_title.split()[:5])
        for _, ref_row in references_df.iterrows():
            ref_title_first_five_words = ' '.join(ref_row['题名'].strip().lower().split()[:5])

            # 将参考文献中的开始页和结束页转换为字符串，同时去掉小数点
            ref_start_page = str(int(float(ref_row['开始页']))) if str(ref_row['开始页']).replace('.', '',
                                                                                                  1).isdigit() else str(
                ref_row['开始页']).strip()
            ref_end_page = str(int(float(ref_row['结束页']))) if str(ref_row['结束页']).replace('.', '',
                                                                                                1).isdigit() else str(
                ref_row['结束页']).strip()
            ref_article_number = str(int(float(ref_row['文献号码']))) if str(ref_row['文献号码']).replace('.', '',
                                                                                                          1).isdigit() else str(
                ref_row['文献号码']).strip()

            if sci_title_first_five_words == ref_title_first_five_words:
                if ((sci_start_page == ref_start_page and sci_end_page == ref_end_page) or
                        (sci_start_page == '' and sci_article_number == ref_article_number)):
                    sci_df.at[i, '论文清单序号'] = ref_row['论文清单序号']
                    matched = True
                    break

        # Fourth, check if start page and end page match
        if not matched:
            for _, ref_row in references_df.iterrows():
                ref_start_page = str(int(float(ref_row['开始页']))) if str(ref_row['开始页']).replace('.', '',
                                                                                                      1).isdigit() else str(
                    ref_row['开始页']).strip()
                ref_end_page = str(int(float(ref_row['结束页']))) if str(ref_row['结束页']).replace('.', '',
                                                                                                    1).isdigit() else str(
                    ref_row['结束页']).strip()
                if (sci_start_page == ref_start_page and sci_end_page == ref_end_page):
                    sci_df.at[i, '论文清单序号'] = ref_row['论文清单序号']
                    matched = True
                    break

        # If no match, set to None
        if not matched:
            sci_df.at[i, '论文清单序号'] = None

    return sci_df


def preserve_formatting_and_export(sci_df, sci_file_path, output_sci_path):
    """
    将带有匹配结果的文献信息导出为 Excel 格式，保留原有的格式，包括单元格内部分文本格式。
    """
    workbook = load_workbook(sci_file_path)
    sheet = workbook.active

    sheet.insert_cols(1)
    sheet.cell(row=1, column=1).value = '论文清单序号'

    arial_font = Font(name='Arial')

    for idx, value in enumerate(sci_df['论文清单序号'], start=2):
        cell = sheet.cell(row=idx, column=1)
        cell.value = value
        cell.font = arial_font

    workbook.save(output_sci_path)
    print(f"更新后的 SCI-E 文件已成功导出至: {output_sci_path}")


def main():
    docx_path = r'examples\张健示例\张健论文清单.docx'
    output_references_path = r'examples\张健示例\委托人论文清单.xlsx'
    sci_file_path = r'examples\张健示例\SCI-E收录数据\SCI-E收录.xlsx'
    output_sci_path = r'data_output\SCI-E收录已标序号.xlsx'

    references = extract_references_from_docx(docx_path)
    export_references_to_excel(references, output_references_path)

    sci_df = pd.read_excel(sci_file_path)
    references_df = pd.read_excel(output_references_path)

    updated_sci_df = match_references(sci_df, references_df)

    preserve_formatting_and_export(updated_sci_df, sci_file_path, output_sci_path)


if __name__ == "__main__":
    main()



def standardize_pinyin_name(name):
    """
    标准化输入拼音名字为 '姓, 名' 的形式
    """
    name = name.replace('-', '').strip()
    name = name.lower()

    if ',' in name:
        parts = [p.strip() for p in name.split(',')]
    else:
        parts = name.split()

    if len(parts) == 2:
        surname = parts[0]
        given_name = parts[1]
    elif len(parts) > 2:
        surname = parts[0]
        given_name = ''.join(parts[1:])
    else:
        return name

    return f"{surname}, {given_name}"


def highlight_names_in_excel(input_path, output_path, author_input):
    df = pd.read_excel(input_path)

    # 按照'论文清单序号'升序排序
    df = df.sort_values(by='论文清单序号')

    standardized_input_name = standardize_pinyin_name(author_input)

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 定义无框线的Arial字体格式
        default_format = workbook.add_format({'font_name': 'Arial', 'border': 0})
        red_bold_format = workbook.add_format({'bold': True, 'font_color': 'red', 'font_name': 'Arial', 'border': 0})

        # 设置整个工作表默认字体为Arial，且无框线
        worksheet.set_default_row(hide_unused_rows=False)
        worksheet.set_column(0, df.shape[1] - 1, None, default_format)

        # 为标题行设置Arial字体和无框线
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, default_format)

        author_col_idx = df.columns.get_loc('Author Full Names')
        for row_idx, author_cell in enumerate(df['Author Full Names'], start=1):
            if pd.notna(author_cell):
                authors = [a.strip() for a in author_cell.split('; ')]
                rich_text_parts = []
                last_pos = 0

                for author in authors:
                    standardized_author = standardize_pinyin_name(author)
                    start_index = author_cell.find(author, last_pos)

                    if standardized_author == standardized_input_name:
                        if start_index > last_pos:
                            # 将前面的部分也设置为Arial字体
                            rich_text_parts.append(default_format)
                            rich_text_parts.append(author_cell[last_pos:start_index])
                        rich_text_parts.append(red_bold_format)
                        rich_text_parts.append(author)
                        last_pos = start_index + len(author)
                    else:
                        if start_index >= 0:
                            rich_text_parts.append(default_format)
                            rich_text_parts.append(author_cell[last_pos:start_index + len(author)])
                            last_pos = start_index + len(author)

                if last_pos < len(author_cell):
                    rich_text_parts.append(default_format)
                    rich_text_parts.append(author_cell[last_pos:])

                rich_text_parts = [part for part in rich_text_parts if part != '']

                if rich_text_parts:
                    worksheet.write_rich_string(row_idx, author_col_idx, *rich_text_parts)
                else:
                    worksheet.write(row_idx, author_col_idx, author_cell, default_format)

        print(f"匹配完成，结果已保存到 {output_path}")


def main():
    input_path = os.path.join('data_output', 'SCI-E收录已标序号.xlsx')
    output_path = os.path.join('data_output', '1_SCI-E收录已标序号已标红.xlsx')

    author_input = input("请输入作者姓名拼音（姓与名中间用逗号或空格分隔）：")

    highlight_names_in_excel(input_path, output_path, author_input)


if __name__ == "__main__":
    main()