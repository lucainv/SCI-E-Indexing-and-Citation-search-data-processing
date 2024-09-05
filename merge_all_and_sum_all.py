import os
from openpyxl import load_workbook, Workbook
from copy import copy
import sys

#汇总引用明细表，3_开头的文件
def merge_excel_files_with_format(folder_path, prefix, output_filename):
    # Get a list of files starting with the given prefix and sort them in ascending order
    file_list = sorted(
        [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.startswith(prefix) and f.endswith('.xlsx')])

    # If no files are found, print a message and exit the program
    if not file_list:
        print(f"没有找到以 {prefix} 开头的文件，合并过程终止。")
        sys.exit(1)

    # Create a new workbook to store the merged result
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = 'Sheet1'

    citation_number = 1  # Initialize citation sequence number

    # Iterate over each file and read its contents
    for idx, file_path in enumerate(file_list):
        wb = load_workbook(file_path)
        ws = wb.active

        # Insert two empty rows before merging the subsequent files
        if idx > 0:  # For files other than the first one
            merged_ws.append(["" for _ in range(ws.max_column)])
            merged_ws.append(["" for _ in range(ws.max_column)])

        # Iterate through each row and copy to the new workbook
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            new_row_idx = merged_ws.max_row + 1
            for cell in row:
                new_cell = merged_ws.cell(row=new_row_idx, column=cell.column, value=cell.value)

                # Copy cell style
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

            # Update citation sequence numbers if needed
            if idx > 0 and row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('被引文献'):
                # Update the citation number without a space and increment it
                merged_ws.cell(row=new_row_idx, column=1, value=f"被引文献{citation_number}")
                citation_number += 1
            elif idx == 0 and row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('被引文献'):
                # Extract the number from the first file and set the citation number
                parts = row[0].value.replace('被引文献', '')
                if parts.isdigit():
                    citation_number = int(parts) + 1

    # Remove the first empty row if it exists
    if all(cell.value is None for cell in merged_ws[1]):
        merged_ws.delete_rows(1)

    # Save the merged file
    output_path = os.path.join(folder_path, output_filename)
    merged_wb.save(output_path)
    print(f"合并完成: {output_path}")


def main():
    # Specify the path to the data_output folder
    folder_path = 'data_output'

    # Merge files starting with 3_
    merge_excel_files_with_format(folder_path, '3_', '3_SCI-E引用明细表_已汇总.xlsx')


if __name__ == "__main__":
    main()



#汇总引用统计表，4_开头的文件
def merge_excel_files_with_continuous_citation_numbers(folder_path, prefix, output_filename):
    # Get a list of files starting with the given prefix and sort them in ascending order
    file_list = sorted(
        [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.startswith(prefix) and f.endswith('.xlsx')])

    # If no files are found, print a message and exit the program
    if not file_list:
        print(f"没有找到以 {prefix} 开头的文件，合并过程终止。")
        sys.exit(1)

    # Initialize a list to hold all data
    all_rows = []

    # Process the first file to include its header
    base_file_path = file_list[0]
    wb = load_workbook(base_file_path)
    ws = wb.active

    # Add all rows from the first file including header
    for row in ws.iter_rows(values_only=False):
        all_rows.append([cell for cell in row])

    # Track the last citation number used
    last_citation_number = 0

    # Process remaining files
    for file_path in file_list[1:]:
        wb = load_workbook(file_path)
        ws = wb.active

        # Skip the header row for subsequent files and only add data rows
        for row in ws.iter_rows(min_row=2, values_only=False):
            new_row = [cell for cell in row]
            all_rows.append(new_row)

    # Update citation numbers in all rows, starting after the header row
    for row in all_rows[1:]:
        if row[0].value and row[0].value.startswith("被引文献"):
            last_citation_number += 1
            row[0].value = f"被引文献{last_citation_number}"

    # Create a new workbook and add rows to it
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = 'Sheet1'

    # Write rows to the new workbook
    for row_index, row in enumerate(all_rows):
        for col_index, cell in enumerate(row):
            new_cell = merged_ws.cell(row=row_index + 1, column=col_index + 1, value=cell.value)

            # Copy cell style
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Remove the first empty row if it exists
    if merged_ws.max_row > 1:
        first_row = merged_ws[1]
        if all(cell.value is None for cell in first_row):
            merged_ws.delete_rows(1)

    # Save the merged file
    output_path = os.path.join(folder_path, output_filename)
    merged_wb.save(output_path)
    print(f"合并完成: {output_path}")


def main():
    # Specify the path to the data_output folder
    folder_path = 'data_output'

    # Merge files starting with 4_
    merge_excel_files_with_continuous_citation_numbers(folder_path, '4_', '4_SCI-E引用统计表_已汇总.xlsx')


if __name__ == "__main__":
    main()




#汇总引用格式for_word表，5_开头的文件
def merge_excel_files_with_sequential_numbers(folder_path, prefix, output_filename):
    # 获取以指定前缀开头的文件列表，并按文件名排序
    file_list = sorted(
        [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.startswith(prefix) and f.endswith('.xlsx')])

    # 如果没有找到匹配的文件，输出提示信息并退出程序
    if not file_list:
        print(f"没有找到以 {prefix} 开头的文件，合并过程终止。")
        sys.exit(1)

    # 初始化用于存储所有行的列表
    all_rows = []

    # 追踪第一列中的最后一个数字
    last_number = None

    # 处理每个文件
    for file_index, file_path in enumerate(file_list):
        wb = load_workbook(file_path)
        ws = wb.active

        # 处理每一行
        for row in ws.iter_rows(min_row=1, values_only=False):
            new_row = [cell for cell in row]

            # 将第一列的值转换为整数形式
            if new_row[0].value is not None:
                try:
                    current_number = int(new_row[0].value)  # 转换为整数
                except ValueError:
                    print(f"无法将 '{new_row[0].value}' 转换为数字，跳过该行")
                    continue

                # 对于第一个文件，保留所有数字不变
                if file_index == 0:
                    last_number = current_number
                else:
                    # 对于后续文件，调整第一列中的数字
                    last_number += 1
                    new_row[0].value = last_number

            all_rows.append(new_row)

    # 创建一个新的工作簿并将行添加到其中
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = 'Sheet1'

    # 将所有行写入新的工作簿
    for row_index, row in enumerate(all_rows):
        for col_index, cell in enumerate(row):
            new_cell = merged_ws.cell(row=row_index + 1, column=col_index + 1, value=cell.value)

            # 复制单元格样式以保持格式
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 如果存在空行，则删除合并后的文件的第一行空行
    if merged_ws.max_row > 1:
        first_row = merged_ws[1]
        if all(cell.value is None for cell in first_row):
            merged_ws.delete_rows(1)

    # 保存合并后的文件
    output_path = os.path.join(folder_path, output_filename)
    merged_wb.save(output_path)
    print(f"合并完成: {output_path}")


def main():
    # 指定data_output文件夹路径
    folder_path = 'data_output'

    # 合并以5_开头的文件
    merge_excel_files_with_sequential_numbers(folder_path, '5_', '5_SCI-E引用格式表_for_word_已汇总.xlsx')


if __name__ == "__main__":
    main()