import re
import os
import pandas as pd
from pypinyin import lazy_pinyin, Style
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from copy import copy


def expand_pinyin_variants(surname, given_name):
    """
    扩展手动输入的拼音形式，生成各种拼音表示形式，包括大写和小写组合
    """
    surname = surname.capitalize().strip()
    given_name = given_name.replace(' ', '').replace('-', '').capitalize().strip()

    variants = [
        f"{surname} {given_name}",
        f"{surname} {given_name.lower()}",
        f"{surname} {given_name.capitalize()}",
        f"{surname}-{given_name.capitalize()}",
        f"{surname}-{given_name.lower()}",
        f"{surname}, {given_name}",
        f"{surname}, {given_name.lower()}",
        f"{surname}, {given_name.capitalize()}",
        f"{surname}, {given_name}",
        f"{surname} {given_name[0].upper() + given_name[1:].lower()}",
        f"{surname} {given_name.lower()}",
        f"{surname} {given_name[0].lower() + given_name[1:].capitalize()}",
        f"{surname}{given_name.lower()}",
        f"{surname}{given_name.capitalize()}",
        f"{surname}-{given_name}",
        f"{surname}-{given_name.lower()}",
        f"{surname}-{given_name.capitalize()}"
    ]

    return variants


def generate_pinyin_variants(chinese_name):
    """
    将中文姓名转换为多种拼音表示形式，包括大写和小写组合
    """
    pinyin_name = lazy_pinyin(chinese_name, style=Style.NORMAL, strict=False)
    if len(pinyin_name) < 2:
        raise ValueError("输入的中文姓名格式不正确，请确保有姓氏和名字。")

    surname = pinyin_name[0].capitalize()
    given_name = ''.join(pinyin_name[1:])
    given_name_hyphenated = '-'.join(pinyin_name[1:])

    variants = [
        f"{surname} {given_name}",
        f"{surname} {given_name.lower()}",
        f"{surname} {given_name.capitalize()}",
        f"{surname}-{given_name.capitalize()}",
        f"{surname}-{given_name.lower()}",
        f"{surname}, {given_name}",
        f"{surname}, {given_name.lower()}",
        f"{surname}, {given_name.capitalize()}",
        f"{surname}, {given_name}",
        f"{surname} {given_name[0].upper() + given_name[1:].lower()}",
        f"{surname} {given_name.lower()}",
        f"{surname} {given_name[0].lower() + given_name[1:].capitalize()}",
        f"{surname}{given_name.lower()}",
        f"{surname}{given_name.capitalize()}",
        f"{surname}-{given_name}",
        f"{surname}-{given_name.lower()}",
        f"{surname}-{given_name.capitalize()}"
    ]

    return variants


def normalize_name(name):
    """
    Normalize name formats to a consistent representation
    """
    # Remove any hyphens and spaces in names
    name = name.replace(' ', '').replace('-', '')
    return name


def parse_manual_pinyin(pinyin_inputs):
    """
    解析手动输入的拼音格式，并生成各种拼音表示形式
    """
    names = []
    for pinyin_input in pinyin_inputs.split(';'):
        parts = re.split(r'[, ]+', pinyin_input.strip())
        if len(parts) >= 2:
            surname = parts[0].strip()
            given_name = ''.join(parts[1:]).strip()
            normalized_given_name = normalize_name(given_name)
            variants = expand_pinyin_variants(surname, normalized_given_name)
            names.extend(variants)
    return names


def highlight_name(input_file, output_file, column_name, names):
    # 读取 .xls 文件并转换为 .xlsx 文件
    intermediate_file = 'intermediate.xlsx'

    # 使用 pandas 读取 .xls 文件
    df = pd.read_excel(input_file, engine='xlrd')

    # 将数据保存为 .xlsx 文件
    df.to_excel(intermediate_file, index=False)

    # 使用 openpyxl 加载 .xlsx 文件
    wb = load_workbook(intermediate_file)
    ws = wb.active

    # 设置黄色填充
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 获取指定列的索引
    col_idx = df.columns.get_loc(column_name) + 1
    author_col_idx = df.columns.get_loc('Authors') + 1  # 获取'Authors'列的索引

    # 统计总数据条数和未被高亮的数据条数
    total_count = 0
    non_highlight_count = 0

    # 遍历指定列，查找包含搜索字符串的单元格并设置高亮
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        cell = row[col_idx - 1]  # 列的索引从0开始，因此需要减去1
        author_cell = row[author_col_idx - 1]  # 同上

        total_count += 1
        cell_values = [value.strip() for value in str(cell.value).split(';')]
        flag = 0
        for name in names:
            if name in cell_values:
                cell.fill = yellow_fill
                author_cell.fill = yellow_fill  # 同时高亮'Authors'字段的单元格
                flag = 1
                break
        if flag == 0:
            non_highlight_count += 1

    # 保存修改后的 Excel 文件
    wb.save(output_file)

    return total_count, non_highlight_count


def highlight_name_batch(input_folder, output_folder, column_name, names):
    total_count_sum = 0
    non_highlight_count_sum = 0

    # 确保输出文件夹存在
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 遍历文件夹中的所有 .xls 文件
    for filename in os.listdir(input_folder):
        if filename.endswith('.xls'):
            input_file = os.path.join(input_folder, filename)
            output_file = os.path.join(output_folder, filename.replace('.xls', '_highlighted.xlsx'))

            # 调用 highlight_name 函数处理文件
            total_count, non_highlight_count = highlight_name(input_file, output_file, column_name, names)

            # 累加统计结果
            total_count_sum += total_count
            non_highlight_count_sum += non_highlight_count

    print(f"总数据条数（总引）: {total_count_sum}")
    print(f"未被高亮的数据条数（他引）: {non_highlight_count_sum}")


if __name__ == '__main__':
    # 示例使用
    input_folder = 'examples/张健示例/SCI-E引用数据/'  # 替换为你的输入文件夹路径
    output_folder = 'data_output/'  # 替换为你的输出文件夹路径
    column_name = 'Author Full Names'  # 替换为你的列名

    # 询问用户是手动输入拼音还是自动生成
    mode = input(
        "请输入需要黄色高亮的自引作者。\n你想手动输入拼音形式还是通过中文姓名自动生成拼音？\n(回复manual手动输入作者姓名拼音，回复auto自动通过作者姓名汉字生成拼音): ").strip().lower()

    if mode == 'manual':
        pinyin_inputs = input(
            "请输入拼音形式：\n（1-姓在前，名在后，全拼；2-姓和名之间可用空格或逗号,连接，请保持一致；3-名有两个汉字的，两个汉字拼音之间用空格或-连接），用';'分隔多个作者的姓名: ").strip()
        # 解析手动输入的拼音形式
        names = parse_manual_pinyin(pinyin_inputs)


    elif mode == 'auto':
        chinese_names = input("请输入中文姓名（多个姓名用逗号、空格、分号或顿号隔开）: ")
        names = []
        for chinese_name in re.split(r'[，、；; ]', chinese_names):
            chinese_name = chinese_name.strip()
            if chinese_name:
                names.extend(generate_pinyin_variants(chinese_name))

    else:
        print("无效的选择，请输入 'manual' 或 'auto'")
        exit(1)

    # 批量处理高亮
    highlight_name_batch(input_folder, output_folder, column_name, names)
    print("处理完成。")